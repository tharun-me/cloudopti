"""Excel export functionality with multiple sheets per service"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


class ExcelExporter:
    """Handle Excel file export with multiple sheets"""
    
    def __init__(self):
        """Initialize exporter"""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14)
        self.bold_font = Font(bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_summary_sheet(self, wb, costs_by_service, total_cost, start_date, end_date, recommendations):
        """Create summary sheet with cost overview"""
        # Get or create summary sheet
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
        else:
            # Use active sheet or create new one
            ws = wb.active
            if ws is not None:
                ws.title = "Summary"
            else:
                ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "AWS Cost Report - Last Month"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:B1')
        
        # Date range
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:B2')
        
        # Empty row
        ws['A3'] = ""
        
        # Cost breakdown header
        ws['A4'] = "Service"
        ws['B4'] = "Cost ($)"
        ws['A4'].fill = self.header_fill
        ws['B4'].fill = self.header_fill
        ws['A4'].font = self.header_font
        ws['B4'].font = self.header_font
        ws['A4'].alignment = Alignment(horizontal='center')
        ws['B4'].alignment = Alignment(horizontal='center')
        ws['A4'].border = self.border
        ws['B4'].border = self.border
        
        # Cost data
        row = 5
        for service, cost in costs_by_service.items():
            ws[f'A{row}'] = service
            ws[f'B{row}'] = round(cost, 2)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = round(total_cost, 2)
        ws[f'A{row}'].font = self.bold_font
        ws[f'B{row}'].font = self.bold_font
        ws[f'A{row}'].border = self.border
        ws[f'B{row}'].border = self.border
        
        # Recommendations section
        row += 2
        ws[f'A{row}'] = "Cost-Saving Recommendations"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        for rec in recommendations:
            ws[f'A{row}'] = f"• {rec}"
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
    
    def create_ec2_sheet(self, wb, ec2_instances, metrics_data, analyses, resource_costs=None):
        """Create EC2 instances sheet"""
        ws = wb.create_sheet("EC2")
        
        # Title
        ws['A1'] = "EC2 Instances - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:O1')
        
        # Headers
        headers = [
            'Instance ID', 'Instance Type', 'Region', 'State', 
            'vCPU', 'Memory (GiB)', 'CPU Avg %', 'CPU Max %',
            'Actual Cost ($)', 'Monitoring', 'CW Agent', 'Recommended Type', 'Potential Savings ($)', 'Savings %', 'Recommendation'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for instance in ec2_instances:
            instance_id = instance['InstanceId']
            instance_type = instance['InstanceType']
            region = instance['Region']
            state = instance['State']
            
            specs = analyses.get(instance_id, {}).get('CurrentSpecs', {})
            vcpu = specs.get('vCPU', 'N/A')
            memory = specs.get('Memory', 'N/A')
            
            metrics = metrics_data.get(instance_id, {})
            cpu_metrics = metrics.get('CPUUtilization', {})
            cpu_avg = cpu_metrics.get('avg', 0) if isinstance(cpu_metrics, dict) else 0
            cpu_max = cpu_metrics.get('max', 0) if isinstance(cpu_metrics, dict) else 0
            
            analysis = analyses.get(instance_id, {})
            # Get actual cost from resource_costs, fallback to calculated cost
            actual_cost = resource_costs.get(instance_id, 0) if resource_costs else analysis.get('CurrentMonthlyCost', 0)
            recommended_type = analysis.get('RecommendedType', 'N/A')
            savings = analysis.get('PotentialSavings', 0)
            savings_percent = analysis.get('SavingsPercent', 0)
            recommendation = analysis.get('Recommendation', 'N/A')
            
            # Get monitoring status
            metrics = metrics_data.get(instance_id, {})
            monitoring = 'Yes' if metrics.get('monitoring_enabled') else 'No'
            cw_agent = 'Yes' if metrics.get('cw_agent_installed') else 'No'
            
            data = [
                instance_id, instance_type, region, state,
                vcpu, memory, round(cpu_avg, 2), round(cpu_max, 2),
                round(actual_cost, 2), monitoring, cw_agent, recommended_type, 
                round(savings, 2), round(savings_percent, 2), recommendation
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 50
    
    def create_eks_sheet(self, wb, eks_clusters, metrics_data):
        """Create EKS clusters sheet"""
        ws = wb.create_sheet("EKS")
        
        # Title
        ws['A1'] = "EKS Clusters - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = [
            'Cluster Name', 'Region', 'Status', 'Version', 'VPC ID',
            'Node Groups', 'CPU Avg %', 'Memory Avg %'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for cluster in eks_clusters:
            cluster_name = cluster['ClusterName']
            region = cluster['Region']
            status = cluster['Status']
            version = cluster['Version']
            vpc_id = cluster['VpcId']
            
            node_groups = cluster.get('NodeGroups', [])
            ng_info = []
            for ng in node_groups:
                ng_info.append(f"{ng.get('Name', 'N/A')} ({ng.get('InstanceType', 'N/A')})")
            node_groups_str = ', '.join(ng_info) if ng_info else 'N/A'
            
            metrics = metrics_data.get(cluster_name, {})
            cpu_avg = metrics.get('CPUUtilization', {}).get('avg', 0) if isinstance(metrics.get('CPUUtilization', {}), dict) else 0
            mem_avg = metrics.get('MemoryUtilization', {}).get('avg', 0) if isinstance(metrics.get('MemoryUtilization', {}), dict) else 0
            
            data = [
                cluster_name, region, status, version, vpc_id,
                node_groups_str, round(cpu_avg, 2), round(mem_avg, 2)
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
    
    def create_s3_sheet(self, wb, s3_buckets, metrics_data):
        """Create S3 buckets sheet"""
        ws = wb.create_sheet("S3")
        
        # Title
        ws['A1'] = "S3 Buckets - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = [
            'Bucket Name', 'Region', 'Size (GB)', 'Object Count',
            'Storage Classes', 'Avg Size (GB)', 'Avg Objects'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for bucket in s3_buckets:
            bucket_name = bucket['BucketName']
            region = bucket['Region']
            size_bytes = bucket.get('Size', 0)
            size_gb = size_bytes / (1024**3)
            object_count = bucket.get('ObjectCount', 0)
            
            storage_classes = bucket.get('StorageClass', {})
            storage_classes_str = ', '.join([f"{k}: {v/(1024**3):.2f}GB" for k, v in storage_classes.items()]) if storage_classes else 'N/A'
            
            metrics = metrics_data.get(bucket_name, {})
            avg_size = metrics.get('BucketSize', {}).get('avg', 0) / (1024**3) if isinstance(metrics.get('BucketSize', {}), dict) else 0
            avg_objects = metrics.get('NumberOfObjects', {}).get('avg', 0) if isinstance(metrics.get('NumberOfObjects', {}), dict) else 0
            
            data = [
                bucket_name, region, round(size_gb, 2), object_count,
                storage_classes_str, round(avg_size, 2), round(avg_objects, 0)
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def create_vpc_sheet(self, wb, vpcs):
        """Create VPCs sheet"""
        ws = wb.create_sheet("VPC")
        
        # Title
        ws['A1'] = "VPCs - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = [
            'VPC ID', 'Region', 'CIDR Block', 'State', 'Is Default', 'Subnet Count'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for vpc in vpcs:
            data = [
                vpc['VpcId'], vpc['Region'], vpc['CidrBlock'],
                vpc['State'], vpc['IsDefault'], vpc.get('SubnetCount', 0)
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def create_rds_sheet(self, wb, rds_instances, resource_costs=None):
        """Create RDS instances sheet"""
        ws = wb.create_sheet("RDS")
        
        # Title
        ws['A1'] = "RDS Instances - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:J1')
        
        # Headers
        headers = [
            'DB Instance ID', 'Instance Class', 'Engine', 'Engine Version',
            'Region', 'Status', 'Multi-AZ', 'Storage Type', 'Storage (GB)', 'Actual Cost ($)', 'VPC ID'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for instance in rds_instances:
            db_id = instance['DBInstanceIdentifier']
            actual_cost = resource_costs.get(db_id, 0) if resource_costs else 0
            
            data = [
                db_id,
                instance['DBInstanceClass'],
                instance['Engine'],
                instance['EngineVersion'],
                instance['Region'],
                instance['Status'],
                instance['MultiAZ'],
                instance['StorageType'],
                instance['AllocatedStorage'],
                round(actual_cost, 2),
                instance['VpcId']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 20
    
    def create_lambda_sheet(self, wb, lambda_functions):
        """Create Lambda functions sheet"""
        ws = wb.create_sheet("Lambda")
        
        # Title
        ws['A1'] = "Lambda Functions - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = [
            'Function Name', 'Runtime', 'Memory (MB)', 'Timeout (s)',
            'Region', 'Code Size (bytes)', 'Last Modified'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for func in lambda_functions:
            data = [
                func['FunctionName'],
                func['Runtime'],
                func['MemorySize'],
                func['Timeout'],
                func['Region'],
                func['CodeSize'],
                func['LastModified'].strftime('%Y-%m-%d') if isinstance(func['LastModified'], datetime) else func['LastModified']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def create_dynamodb_sheet(self, wb, dynamodb_tables):
        """Create DynamoDB tables sheet"""
        ws = wb.create_sheet("DynamoDB")
        
        # Title
        ws['A1'] = "DynamoDB Tables - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = [
            'Table Name', 'Region', 'Status', 'Item Count',
            'Table Size (bytes)', 'Billing Mode'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for table in dynamodb_tables:
            data = [
                table['TableName'],
                table['Region'],
                table['Status'],
                table['ItemCount'],
                table['TableSizeBytes'],
                table['BillingMode']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
    
    def create_elb_sheet(self, wb, load_balancers):
        """Create Load Balancers sheet"""
        ws = wb.create_sheet("ELB")
        
        # Title
        ws['A1'] = "Load Balancers - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = [
            'Load Balancer Name', 'Type', 'Scheme', 'Region', 'State', 'VPC ID'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for lb in load_balancers:
            data = [
                lb['LoadBalancerName'],
                lb['Type'],
                lb['Scheme'],
                lb['Region'],
                lb['State'],
                lb['VpcId']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
    
    def create_route53_sheet(self, wb, hosted_zones):
        """Create Route53 hosted zones sheet"""
        ws = wb.create_sheet("Route53")
        
        # Title
        ws['A1'] = "Route53 Hosted Zones - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = [
            'Hosted Zone ID', 'Name', 'Private Zone', 'Record Count'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for zone in hosted_zones:
            data = [
                zone['HostedZoneId'],
                zone['Name'],
                zone['PrivateZone'],
                zone['ResourceRecordSetCount']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
    
    def create_secrets_manager_sheet(self, wb, secrets):
        """Create Secrets Manager secrets sheet"""
        ws = wb.create_sheet("SecretsManager")
        
        # Title
        ws['A1'] = "Secrets Manager Secrets - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = [
            'Secret Name', 'ARN', 'Region', 'Description', 'Last Changed', 'Last Accessed'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for secret in secrets:
            data = [
                secret['SecretName'],
                secret['ARN'],
                secret['Region'],
                secret['Description'],
                secret['LastChangedDate'].strftime('%Y-%m-%d') if isinstance(secret['LastChangedDate'], datetime) else secret['LastChangedDate'],
                secret['LastAccessedDate'].strftime('%Y-%m-%d') if isinstance(secret['LastAccessedDate'], datetime) else secret['LastAccessedDate']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def create_systems_manager_sheet(self, wb, instances):
        """Create Systems Manager instances sheet"""
        ws = wb.create_sheet("SystemsManager")
        
        # Title
        ws['A1'] = "Systems Manager Instances - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = [
            'Instance ID', 'Computer Name', 'Platform Type', 'Platform Version', 'Region', 'Ping Status', 'Last Ping'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for instance in instances:
            data = [
                instance['InstanceId'],
                instance['ComputerName'],
                instance['PlatformType'],
                instance['PlatformVersion'],
                instance['Region'],
                instance['PingStatus'],
                instance['LastPingDateTime'].strftime('%Y-%m-%d') if isinstance(instance['LastPingDateTime'], datetime) else instance['LastPingDateTime']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
    
    def create_cloudwatch_sheet(self, wb, alarms):
        """Create CloudWatch alarms sheet"""
        ws = wb.create_sheet("CloudWatch")
        
        # Title
        ws['A1'] = "CloudWatch Alarms - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = [
            'Alarm Name', 'Namespace', 'Metric Name', 'State', 'Region', 'Last Updated'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for alarm in alarms:
            data = [
                alarm['AlarmName'],
                alarm['Namespace'],
                alarm['MetricName'],
                alarm['StateValue'],
                alarm['Region'],
                alarm['StateUpdatedTimestamp'].strftime('%Y-%m-%d %H:%M') if isinstance(alarm['StateUpdatedTimestamp'], datetime) else alarm['StateUpdatedTimestamp']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
    
    def create_guardduty_sheet(self, wb, detectors):
        """Create GuardDuty detectors sheet"""
        ws = wb.create_sheet("GuardDuty")
        
        # Title
        ws['A1'] = "GuardDuty Detectors - Detailed Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = [
            'Detector ID', 'Region', 'Status', 'Created At', 'Finding Publishing Frequency'
        ]
        
        row = 3
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
            col += 1
        
        # Data rows
        row = 4
        for detector in detectors:
            data = [
                detector['DetectorId'],
                detector['Region'],
                detector['Status'],
                detector['CreatedAt'].strftime('%Y-%m-%d') if isinstance(detector['CreatedAt'], datetime) else detector['CreatedAt'],
                detector['FindingPublishingFrequency']
            ]
            
            col = 1
            for value in data:
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                col += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
    
    def export(self, costs_by_service, total_cost, start_date, end_date, recommendations,
               resources=None, metrics_data=None, analyses=None, resource_costs=None):
        """Export comprehensive report to Excel file"""
        wb = Workbook()
        
        # Create summary sheet first (use the default sheet)
        ws = wb.active
        ws.title = "Summary"
        
        # Create summary sheet content
        self.create_summary_sheet(wb, costs_by_service, total_cost, start_date, end_date, recommendations)
        
        # Create service-specific sheets if resources are provided
        if resources:
            metrics_data = metrics_data or {}
            analyses = analyses or {}
            
            # EC2 sheet
            if resources.get('EC2'):
                self.create_ec2_sheet(wb, resources['EC2'], metrics_data, analyses)
            
            # EKS sheet
            if resources.get('EKS'):
                self.create_eks_sheet(wb, resources['EKS'], metrics_data)
            
            # S3 sheet
            if resources.get('S3'):
                self.create_s3_sheet(wb, resources['S3'], metrics_data)
            
            # VPC sheet
            if resources.get('VPC'):
                self.create_vpc_sheet(wb, resources['VPC'])
            
            # RDS sheet
            if resources.get('RDS'):
                self.create_rds_sheet(wb, resources['RDS'], resource_costs.get('RDS', {}) if resource_costs else {})
            
            # Lambda sheet
            if resources.get('Lambda'):
                self.create_lambda_sheet(wb, resources['Lambda'])
            
            # DynamoDB sheet
            if resources.get('DynamoDB'):
                self.create_dynamodb_sheet(wb, resources['DynamoDB'])
            
            # ELB sheet
            if resources.get('ELB'):
                self.create_elb_sheet(wb, resources['ELB'])
            
            # Route53 sheet
            if resources.get('Route53'):
                self.create_route53_sheet(wb, resources['Route53'])
            
            # SecretsManager sheet
            if resources.get('SecretsManager'):
                self.create_secrets_manager_sheet(wb, resources['SecretsManager'])
            
            # SystemsManager sheet
            if resources.get('SystemsManager'):
                self.create_systems_manager_sheet(wb, resources['SystemsManager'])
            
            # CloudWatch sheet
            if resources.get('CloudWatch'):
                self.create_cloudwatch_sheet(wb, resources['CloudWatch'])
            
            # GuardDuty sheet
            if resources.get('GuardDuty'):
                self.create_guardduty_sheet(wb, resources['GuardDuty'])
        
        # Save file
        filename = f"aws_cost_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
